"""Excel写入模块 - 支持模板填充"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path
from datetime import datetime
import yaml


def load_config():
    """加载配置"""
    config_path = Path(__file__).parent.parent / "config.yaml"
    if not config_path.exists():
        config_path = Path(__file__).parent.parent / "config.example.yaml"
    
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def write_to_excel(df, excel_path, sheet_name='一览表', start_col='E', start_row=4, 
                   date_cell='B1', exec_date=None):
    """
    写入数据到现有Excel
    
    参数:
        df: DataFrame数据
        excel_path: Excel文件路径
        sheet_name: Sheet名称
        start_col: 数据起始列 (默认E)
        start_row: 数据起始行 (默认4)
        date_cell: 日期写入单元格 (默认B1)
        exec_date: 执行日期时间
    """
    exec_date = exec_date or datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # 加载现有Excel
    wb = load_workbook(excel_path)
    
    # 获取或创建Sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
        ws.title = sheet_name
    
    # 写入执行时间到B1
    ws['B1'] = exec_date
    ws['B1'].font = Font(bold=True, size=12)
    
    # 将列字母转为数字 (A=1, B=2, E=5, ...)
    start_col_num = ord(start_col.upper()) - ord('A') + 1
    
    # 清空现有数据区域 (从起始位置开始)
    # 计算数据占据的列范围
    max_col = start_col_num + len(df.columns) - 1
    for row in range(start_row, ws.max_row + 1):
        for col in range(start_col_num, max_col + 1):
            ws.cell(row=row, column=col).value = None
    
    # 写入表头 (如果从第4行开始，通常保留原有表头)
    # 只写入数据，不写表头
    
    # 写入数据
    for r_idx, row_data in enumerate(df.values, start=start_row):
        for c_idx, value in enumerate(row_data, start=start_col_num):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value
            
            # 格式化数值
            if pd.notna(value):
                if isinstance(value, float):
                    # 判断是否为百分比 (增长率列)
                    col_name = df.columns[c_idx - start_col_num]
                    if '增长率' in col_name:
                        cell.number_format = '0.00%'
                    else:
                        cell.number_format = '#,##0.00'
                elif isinstance(value, int):
                    cell.number_format = '#,##0'
    
    # 保存
    wb.save(excel_path)
    print(f"数据已写入: {excel_path}")
    return excel_path


def write_simple(df, output_path, sheet_name='数据'):
    """
    直接写入Excel（无模板）
    """
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # 美化表格
    style_excel(output_path, sheet_name)


def style_excel(output_path, sheet_name='数据'):
    """美化Excel表格"""
    wb = load_workbook(output_path)
    ws = wb[sheet_name]
    
    # 标题样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置标题行
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置数据行
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)


def write_output(result_data, output_filename=None):
    """
    生成输出文件 (用于无模板情况)
    result_data: 字典，key为脚本名（不含.sql），value为DataFrame
    """
    config = load_config()
    output_dir = Path(config['excel']['output_dir'])
    output_dir.mkdir(parents=True, exist_ok=True)
    
    if output_filename is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f'report_{timestamp}.xlsx'
    
    output_path = output_dir / output_filename
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, df in result_data.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    
    return output_path
